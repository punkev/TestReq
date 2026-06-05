#!/usr/bin/env python3
"""
secret_finder.py

Recursive hardcoded secret scanner for Java Spring / Struts codebases.

Output:
  - Excel workbook (.xlsx) with one finding per row
  - Optional CSV file

Install:
  pip install openpyxl

Usage:
  python secret_finder.py /path/to/repo -o hardcoded_secrets.xlsx
  python secret_finder.py /path/to/repo -o hardcoded_secrets.xlsx --include-tests
  python secret_finder.py /path/to/repo -o hardcoded_secrets.xlsx --csv hardcoded_secrets.csv

Notes:
  - This is an offline regex + entropy + context scanner.
  - It does not validate whether a credential is live.
  - Treat findings as leads for review, rotation, and removal from source/history.
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import math
import os
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Iterator, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# -----------------------------
# Tuning: repository traversal
# -----------------------------

DEFAULT_EXCLUDED_DIRS = {
    ".git", ".svn", ".hg", ".idea", ".vscode", ".settings",
    "node_modules", "bower_components", "vendor", "vendors",
    "target", "build", "dist", "out", "bin", "obj", ".gradle", ".mvn",
    "coverage", ".nyc_output", ".pytest_cache", "__pycache__",
    "logs", "log", "tmp", "temp", ".terraform", ".serverless",
}

TEST_DIR_NAMES = {
    "test", "tests", "testing", "spec", "specs", "mock", "mocks",
    "fixture", "fixtures", "sample", "samples", "example", "examples",
    "src/test", "src\\test",
}

DEFAULT_ALLOWED_EXTENSIONS = {
    # Java / JVM / Spring / Struts
    ".java", ".kt", ".groovy", ".scala",
    ".properties", ".yml", ".yaml", ".xml", ".jsp", ".jspx", ".tag", ".tld",
    ".gradle", ".pom", ".conf", ".cfg", ".ini", ".env",
    # common web/config files seen in Java repos
    ".json", ".js", ".ts", ".html", ".htm", ".sh", ".bash", ".zsh",
    ".bat", ".cmd", ".dockerfile", ".tf", ".tfvars",
    # no-extension files are handled separately by name
}

ALLOWED_FILENAMES = {
    "dockerfile", "makefile", "jenkinsfile", ".env", ".env.local", ".env.dev",
    ".env.development", ".env.prod", ".env.production", "application",
}

DEFAULT_EXCLUDED_EXTENSIONS = {
    ".class", ".jar", ".war", ".ear", ".zip", ".tar", ".gz", ".tgz", ".7z", ".rar",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico", ".svg", ".pdf",
    ".mp4", ".mov", ".avi", ".mp3", ".wav", ".woff", ".woff2", ".ttf", ".eot",
    ".lock", ".sum",
}

DEFAULT_EXCLUDED_FILE_REGEXES = [
    re.compile(r"(^|[/\\])package-lock\.json$", re.I),
    re.compile(r"(^|[/\\])yarn\.lock$", re.I),
    re.compile(r"(^|[/\\])pnpm-lock\.yaml$", re.I),
    re.compile(r"(^|[/\\])gradle\.lockfile$", re.I),
    re.compile(r"(^|[/\\])dependency-reduced-pom\.xml$", re.I),
]

TEST_FILE_REGEXES = [
    re.compile(r"(^|[/\\])src[/\\]test[/\\]", re.I),
    re.compile(r"(^|[/\\])(test|tests|testing|spec|specs|mock|mocks|fixture|fixtures|sample|samples|example|examples)([/\\]|$)", re.I),
    re.compile(r"(Test|Tests|IT|ITCase|Spec)\.(java|kt|groovy|scala)$", re.I),
    re.compile(r"(mock|fixture|sample|example|dummy|fake)", re.I),
]

MAX_DEFAULT_FILE_SIZE_BYTES = 2_000_000


# -----------------------------
# Tuning: false-positive control
# -----------------------------

DUMMY_VALUES = {
    "", "null", "none", "nil", "undefined", "true", "false",
    "password", "passwd", "pwd", "secret", "token", "changeme", "changeit",
    "example", "examplepassword", "sample", "dummy", "fake", "test", "testing",
    "admin", "root", "user", "username", "guest", "demo", "local", "localhost",
    "your_password", "your-password", "your_secret", "your-secret", "your_token", "your-token",
    "xxxxxxxx", "xxxx", "abc123", "123456", "12345678", "123456789", "qwerty",
}

DUMMY_VALUE_REGEXES = [
    re.compile(r"^\$\{[^}]+\}$"),               # ${ENV_VAR}
    re.compile(r"^%[A-Z0-9_]+%$"),               # %ENV_VAR%
    re.compile(r"^<[^>]+>$"),                    # <password>
    re.compile(r"^\{\{[^}]+\}\}$"),             # {{ secret }}
    re.compile(r"^\$[A-Z0-9_]+$", re.I),         # $ENV
    re.compile(r"^\*+$"),
    re.compile(r"^x+$", re.I),
    re.compile(r"^0+$"),
    re.compile(r"^1+$"),
]

SECRET_CONTEXT_WORDS = re.compile(
    r"(?i)(secret|password|passwd|pwd|token|api[_-]?key|apikey|access[_-]?key|"
    r"private[_-]?key|client[_-]?secret|consumer[_-]?secret|bearer|authorization|"
    r"credential|creds|auth|session|jwt|refresh[_-]?token|db[_-]?password|"
    r"datasource|jdbc|spring\.datasource|hibernate|connection|string|"
    r"apigee|apim|api[_-]?management|subscription[_-]?key|ocp-apim-subscription-key|"
    r"coin|crypto|wallet|private|seed|mnemonic|rpc|web3|etherscan|infura|alchemy)"
)

ASSIGNMENT_CHARS = r"(?:=|:|:=|=>|->)"
QUOTE_CHARS = r"[\"']?"
VALUE_CHARS = r"([^\s\"'`,;#<>]{6,})"


@dataclass(frozen=True)
class Rule:
    name: str
    regex: re.Pattern
    severity: str
    confidence: str
    description: str
    secret_group: str = "secret"
    requires_entropy: bool = False
    min_entropy: float = 0.0
    tags: tuple[str, ...] = ()


@dataclass
class Finding:
    rule_name: str
    severity: str
    confidence: str
    file: str
    line: int
    column: int
    matched_text: str
    secret_masked: str
    secret_sha256_12: str
    entropy: float
    likely_dummy: bool
    context: str
    description: str
    tags: str


# -----------------------------
# Secret regex rules
# -----------------------------

RULES: list[Rule] = [
    # Private keys and certificates
    Rule("Private key block", re.compile(r"(?P<secret>-----BEGIN (?:RSA |DSA |EC |OPENSSH |PGP )?PRIVATE KEY-----)", re.I), "Critical", "High", "Private key material starts here", tags=("private-key",)),
    Rule("Generic PEM key/cert block", re.compile(r"(?P<secret>-----BEGIN [A-Z0-9 ]{0,40}(?:KEY|CERTIFICATE)-----)", re.I), "High", "Medium", "PEM key or certificate block starts here", tags=("pem",)),

    # Cloud/provider-specific patterns
    Rule("AWS access key id", re.compile(r"(?P<secret>\b(?:AKIA|ASIA)[0-9A-Z]{16}\b)"), "Critical", "High", "AWS access key ID", tags=("aws",)),
    Rule("AWS secret access key assignment", re.compile(r"(?i)(?:aws.{0,20})?(?:secret[_-]?access[_-]?key|aws_secret_access_key)\s*" + ASSIGNMENT_CHARS + r"\s*" + QUOTE_CHARS + r"(?P<secret>[A-Za-z0-9/+=]{40})"), "Critical", "High", "AWS secret access key assigned in code/config", tags=("aws",)),
    Rule("Google API key", re.compile(r"(?P<secret>\bAIza[0-9A-Za-z_\-]{35}\b)"), "High", "High", "Google API key", tags=("google",)),
    Rule("Google OAuth client secret", re.compile(r"(?P<secret>\bGOCSPX-[0-9A-Za-z_\-]{20,80}\b)"), "High", "High", "Google OAuth client secret", tags=("google", "oauth")),
    Rule("GCP service account private key id", re.compile(r"(?i)\"private_key_id\"\s*:\s*\"(?P<secret>[a-f0-9]{40})\""), "High", "High", "GCP service account private_key_id", tags=("gcp",)),
    Rule("Azure storage account key", re.compile(r"(?i)(?:account[_-]?key|azure.{0,20}key)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[A-Za-z0-9+/]{80,}={0,2})"), "Critical", "Medium", "Azure-style base64 storage/account key", True, min_entropy=4.5, tags=("azure",)),
    Rule("Azure API Management subscription key header", re.compile(r"(?i)(?:ocp-apim-subscription-key|apim[_-]?subscription[_-]?key|subscription[_-]?key)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[a-f0-9]{32}|[A-Za-z0-9_\-]{24,80})"), "High", "High", "Azure APIM subscription key or header value", tags=("apim", "azure")),

    # Git / SaaS tokens
    Rule("GitHub classic token", re.compile(r"(?P<secret>\bghp_[A-Za-z0-9]{36}\b)"), "Critical", "High", "GitHub personal access token", tags=("github",)),
    Rule("GitHub fine-grained or app token", re.compile(r"(?P<secret>\b(?:github_pat|gho|ghu|ghs|ghr)_[A-Za-z0-9_]{20,255}\b)"), "Critical", "High", "GitHub token", tags=("github",)),
    Rule("GitLab token", re.compile(r"(?P<secret>\bglpat-[A-Za-z0-9_\-]{20}\b)"), "Critical", "High", "GitLab personal access token", tags=("gitlab",)),
    Rule("Slack token", re.compile(r"(?P<secret>\bxox[baprs]-[A-Za-z0-9\-]{10,120}\b)"), "Critical", "High", "Slack token", tags=("slack",)),
    Rule("Stripe secret key", re.compile(r"(?P<secret>\bsk_(?:live|test)_[A-Za-z0-9]{16,120}\b)"), "Critical", "High", "Stripe secret key", tags=("stripe",)),
    Rule("SendGrid API key", re.compile(r"(?P<secret>\bSG\.[A-Za-z0-9_\-]{16,}\.[A-Za-z0-9_\-]{16,}\b)"), "Critical", "High", "SendGrid API key", tags=("sendgrid",)),
    Rule("Twilio API key", re.compile(r"(?P<secret>\bSK[0-9a-fA-F]{32}\b)"), "High", "High", "Twilio-style API key", tags=("twilio",)),
    Rule("Heroku API key", re.compile(r"(?i)(?:heroku.{0,20}(?:api[_-]?key|token)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?)(?P<secret>[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})"), "High", "High", "Heroku API key assigned in code/config", tags=("heroku",)),
    Rule("NPM token", re.compile(r"(?P<secret>\bnpm_[A-Za-z0-9]{36}\b)"), "High", "High", "NPM token", tags=("npm",)),
    Rule("PyPI token", re.compile(r"(?P<secret>\bpypi-[A-Za-z0-9_\-]{50,}\b)"), "High", "High", "PyPI token", tags=("pypi",)),
    Rule("OpenAI API key", re.compile(r"(?P<secret>\bsk-(?:proj-)?[A-Za-z0-9_\-]{20,200}\b)"), "High", "Medium", "OpenAI-style API key", tags=("openai",)),
    Rule("JWT token", re.compile(r"(?P<secret>\beyJ[A-Za-z0-9_\-]{10,}\.[A-Za-z0-9_\-]{10,}\.[A-Za-z0-9_\-]{10,}\b)"), "High", "High", "JWT token", tags=("jwt",)),

    # Database / connection strings
    Rule("JDBC URL with credentials", re.compile(r"(?i)(?P<secret>jdbc:[a-z0-9]+:[^\s\"'<>]*(?:user|username|password|pwd)=[^\s\"'<>]+)"), "Critical", "High", "JDBC URL containing credential parameters", tags=("jdbc", "database", "java")),
    Rule("Database URI with credentials", re.compile(r"(?i)(?P<secret>\b(?:postgres(?:ql)?|mysql|mariadb|mongodb(?:\+srv)?|redis|amqp|rabbitmq)://[^\s:/@]+:[^\s@/]+@[^\s\"'<>]+)"), "Critical", "High", "Database/message-broker URI with embedded username and password", tags=("database", "uri")),
    Rule("Spring datasource password", re.compile(r"(?i)(?:spring\.datasource\.(?:password|hikari\.password)|datasource\.password|db\.password|database\.password)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[^\s\"'#,;]{4,})"), "Critical", "High", "Spring/JDBC datasource password", tags=("spring", "database", "java")),
    Rule("Hibernate connection password", re.compile(r"(?i)(?:hibernate\.connection\.password|javax\.persistence\.jdbc\.password)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[^\s\"'#,;]{4,})"), "Critical", "High", "Hibernate/JPA DB password", tags=("hibernate", "database", "java")),
    Rule("XML password property", re.compile(r"(?i)<property\s+name=[\"'][^\"']*(?:password|secret|token|apiKey|apikey)[^\"']*[\"']\s+value=[\"'](?P<secret>[^\"']{4,})[\"']"), "High", "High", "XML/Spring bean property contains hardcoded secret", tags=("xml", "spring", "struts")),

    # Apigee / API Management / gateway-specific
    Rule("Apigee client secret or consumer secret", re.compile(r"(?i)(?:apigee[^\n]{0,80})?(?:client[_-]?secret|consumer[_-]?secret|app[_-]?secret)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[A-Za-z0-9_\-./+=]{8,200})"), "Critical", "Medium", "Apigee client/consumer secret assignment", True, min_entropy=3.2, tags=("apigee", "api-management")),
    Rule("Apigee management password", re.compile(r"(?i)(?:apigee[^\n]{0,80})?(?:mgmt|management|edge)[_.-]?(?:password|pwd)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[^\s\"'#,;]{6,})"), "Critical", "High", "Apigee management/Edge password", tags=("apigee", "api-management")),
    Rule("API gateway secret", re.compile(r"(?i)(?:api[_-]?gateway|apim|api[_-]?management|gateway)[^\n]{0,60}(?:secret|token|key|password)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[A-Za-z0-9_\-./+=]{8,200})"), "High", "Medium", "API gateway / API management secret", True, min_entropy=3.2, tags=("api-management",)),

    # Crypto / coin / blockchain secrets
    Rule("Ethereum private key assignment", re.compile(r"(?i)(?:private[_-]?key|wallet[_-]?key|eth[_-]?private[_-]?key)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>0x[a-f0-9]{64}|[a-f0-9]{64})\b"), "Critical", "High", "Ethereum/blockchain private key", tags=("crypto", "coin", "wallet")),
    Rule("Crypto seed phrase assignment", re.compile(r"(?i)(?:seed[_-]?phrase|mnemonic|recovery[_-]?phrase)\s*" + ASSIGNMENT_CHARS + r"\s*[\"'](?P<secret>(?:[a-z]+\s+){11,23}[a-z]+)[\"']"), "Critical", "Medium", "Possible wallet seed or mnemonic phrase", tags=("crypto", "wallet", "seed")),
    Rule("Infura API key/url", re.compile(r"(?P<secret>https://[A-Za-z0-9\-.]*infura\.io/v3/[A-Za-z0-9]{20,80})"), "High", "High", "Infura project URL/key", tags=("crypto", "infura")),
    Rule("Alchemy API key/url", re.compile(r"(?P<secret>https://[A-Za-z0-9\-.]*g\.alchemy\.com/v2/[A-Za-z0-9_\-]{20,100})"), "High", "High", "Alchemy API URL/key", tags=("crypto", "alchemy")),

    # Generic high-signal assignments
    Rule("Generic secret assignment", re.compile(r"(?i)(?:^|[\s\"'])((?:[a-z0-9_.-]{0,40})?(?:secret|client_secret|consumer_secret|api_key|apikey|access_key|private_key|auth_token|access_token|refresh_token|bearer_token|session_token|subscription_key|signing_key|encryption_key)(?:[a-z0-9_.-]{0,40})?)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[A-Za-z0-9_\-./+=:]{8,300})"), "High", "Medium", "Generic secret/token/key assignment", True, min_entropy=3.0, tags=("generic",)),
    Rule("Generic password assignment", re.compile(r"(?i)(?:^|[\s\"'])([a-z0-9_.-]{0,40}(?:password|passwd|pwd)[a-z0-9_.-]{0,40})\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[^\s\"'`,;#<>]{4,200})"), "High", "Medium", "Generic hardcoded password assignment", tags=("generic", "password")),
    Rule("Authorization bearer literal", re.compile(r"(?i)(?:authorization|bearer)\s*" + ASSIGNMENT_CHARS + r"?\s*[\"']?Bearer\s+(?P<secret>[A-Za-z0-9_\-./+=]{10,300})"), "High", "Medium", "Hardcoded Authorization Bearer token", True, min_entropy=3.0, tags=("generic", "bearer")),
    Rule("Basic auth URL", re.compile(r"(?i)(?P<secret>https?://[^\s:/@]+:[^\s@/]+@[^\s\"'<>]+)"), "High", "High", "URL containing embedded basic-auth credentials", tags=("url", "basic-auth")),

    # Java keystore / encryption / signing material
    Rule("Java keystore password", re.compile(r"(?i)(?:key-store-password|keystore\.password|truststore\.password|javax\.net\.ssl\.(?:keyStorePassword|trustStorePassword))\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[^\s\"'#,;]{4,})"), "Critical", "High", "Java keystore/truststore password", tags=("java", "keystore", "tls")),
    Rule("Jasypt encryptor password", re.compile(r"(?i)(?:jasypt\.encryptor\.password|encryptor\.password|encryption\.password|encryption\.key)\s*" + ASSIGNMENT_CHARS + r"\s*[\"']?(?P<secret>[^\s\"'#,;]{4,})"), "Critical", "High", "Jasypt/encryption password or key", tags=("java", "spring", "crypto")),
]


# -----------------------------
# Utility functions
# -----------------------------

def shannon_entropy(value: str) -> float:
    if not value:
        return 0.0
    counts = {}
    for ch in value:
        counts[ch] = counts.get(ch, 0) + 1
    entropy = 0.0
    length = len(value)
    for count in counts.values():
        p = count / length
        entropy -= p * math.log2(p)
    return round(entropy, 3)


def mask_secret(secret: str) -> str:
    if len(secret) <= 8:
        return "*" * len(secret)
    if len(secret) <= 16:
        return secret[:2] + "*" * (len(secret) - 4) + secret[-2:]
    return secret[:4] + "*" * min(24, len(secret) - 8) + secret[-4:]


def sha256_12(secret: str) -> str:
    return hashlib.sha256(secret.encode("utf-8", errors="ignore")).hexdigest()[:12]


def normalize_secret(secret: str) -> str:
    return secret.strip().strip('"\'`').strip()


def is_likely_dummy(secret: str) -> bool:
    s = normalize_secret(secret)
    lower = s.lower()
    if lower in DUMMY_VALUES:
        return True
    if len(s) < 4:
        return True
    for rx in DUMMY_VALUE_REGEXES:
        if rx.match(s):
            return True
    # repeated same short pattern, such as abcabcabc or 123123123
    if len(set(s)) <= 2 and len(s) >= 6:
        return True
    return False


def is_binary_file(path: Path, sample_size: int = 4096) -> bool:
    try:
        with path.open("rb") as f:
            sample = f.read(sample_size)
        if b"\x00" in sample:
            return True
        return False
    except OSError:
        return True


def should_skip_file(
    path: Path,
    root: Path,
    include_tests: bool,
    max_file_size: int,
    extra_excluded_dirs: set[str],
    extra_included_exts: set[str],
) -> tuple[bool, str]:
    try:
        rel = path.relative_to(root)
    except ValueError:
        rel = path
    rel_str = str(rel)
    lower_rel = rel_str.lower()

    parts_lower = {p.lower() for p in rel.parts}
    excluded_dirs = {d.lower() for d in DEFAULT_EXCLUDED_DIRS | extra_excluded_dirs}
    if parts_lower & excluded_dirs:
        return True, "excluded directory"

    if not include_tests:
        for rx in TEST_FILE_REGEXES:
            if rx.search(rel_str):
                return True, "test/sample/mock path"

    for rx in DEFAULT_EXCLUDED_FILE_REGEXES:
        if rx.search(rel_str):
            return True, "excluded file pattern"

    suffix = path.suffix.lower()
    if suffix in DEFAULT_EXCLUDED_EXTENSIONS:
        return True, "excluded extension"

    try:
        if path.stat().st_size > max_file_size:
            return True, "file too large"
    except OSError:
        return True, "cannot stat file"

    name_lower = path.name.lower()
    allowed = suffix in (DEFAULT_ALLOWED_EXTENSIONS | extra_included_exts) or name_lower in ALLOWED_FILENAMES
    if not allowed:
        return True, "extension/name not in scan list"

    if is_binary_file(path):
        return True, "binary file"

    return False, ""


def iter_files(
    root: Path,
    include_tests: bool,
    max_file_size: int,
    extra_excluded_dirs: set[str],
    extra_included_exts: set[str],
) -> Iterator[Path]:
    for current_root, dirs, files in os.walk(root):
        current = Path(current_root)
        # prune excluded dirs early
        dirs[:] = [
            d for d in dirs
            if d.lower() not in {x.lower() for x in DEFAULT_EXCLUDED_DIRS | extra_excluded_dirs}
        ]
        for filename in files:
            path = current / filename
            skip, _reason = should_skip_file(path, root, include_tests, max_file_size, extra_excluded_dirs, extra_included_exts)
            if not skip:
                yield path


def read_text_safely(path: Path) -> Optional[str]:
    encodings = ["utf-8", "utf-8-sig", "latin-1"]
    for enc in encodings:
        try:
            return path.read_text(encoding=enc, errors="replace")
        except OSError:
            return None
    return None


def line_col_from_offset(text: str, offset: int) -> tuple[int, int]:
    line = text.count("\n", 0, offset) + 1
    line_start = text.rfind("\n", 0, offset) + 1
    col = offset - line_start + 1
    return line, col


def get_line(text: str, line_no: int) -> str:
    lines = text.splitlines()
    if 1 <= line_no <= len(lines):
        return lines[line_no - 1].strip()
    return ""


def context_score(line: str, secret: str) -> int:
    score = 0
    if SECRET_CONTEXT_WORDS.search(line):
        score += 2
    ent = shannon_entropy(secret)
    if ent >= 4.0:
        score += 2
    elif ent >= 3.0:
        score += 1
    if len(secret) >= 20:
        score += 1
    if re.search(r"[A-Z]", secret) and re.search(r"[a-z]", secret) and re.search(r"\d", secret):
        score += 1
    return score


def adjust_confidence(rule: Rule, secret: str, line: str, dummy: bool) -> str:
    if dummy:
        return "Low"
    score = context_score(line, secret)
    if rule.confidence == "High":
        return "High" if score >= 1 or len(secret) >= 16 else "Medium"
    if rule.confidence == "Medium":
        if score >= 4:
            return "High"
        if score >= 2:
            return "Medium"
        return "Low"
    return rule.confidence


def extract_secret(match: re.Match, rule: Rule) -> str:
    try:
        return normalize_secret(match.group(rule.secret_group))
    except IndexError:
        return normalize_secret(match.group(0))


def scan_text(path: Path, root: Path, text: str) -> list[Finding]:
    findings: list[Finding] = []
    seen: set[tuple[str, str, int]] = set()
    rel_file = str(path.relative_to(root)) if path.is_relative_to(root) else str(path)

    for rule in RULES:
        for match in rule.regex.finditer(text):
            secret = extract_secret(match, rule)
            if not secret:
                continue
            ent = shannon_entropy(secret)
            if rule.requires_entropy and ent < rule.min_entropy:
                continue
            dummy = is_likely_dummy(secret)

            line_no, col = line_col_from_offset(text, match.start())
            line_text = get_line(text, line_no)

            # Do not completely discard dummy-looking secrets; report with Low confidence.
            confidence = adjust_confidence(rule, secret, line_text, dummy)
            key = (rule.name, sha256_12(secret), line_no)
            if key in seen:
                continue
            seen.add(key)

            findings.append(Finding(
                rule_name=rule.name,
                severity=rule.severity,
                confidence=confidence,
                file=rel_file,
                line=line_no,
                column=col,
                matched_text=match.group(0)[:500],
                secret_masked=mask_secret(secret),
                secret_sha256_12=sha256_12(secret),
                entropy=ent,
                likely_dummy=dummy,
                context=line_text[:1000],
                description=rule.description,
                tags=",".join(rule.tags),
            ))
    return findings


def sort_findings(findings: list[Finding]) -> list[Finding]:
    sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    conf_order = {"High": 0, "Medium": 1, "Low": 2}
    return sorted(findings, key=lambda f: (sev_order.get(f.severity, 99), conf_order.get(f.confidence, 99), f.file, f.line))


def write_excel(findings: list[Finding], out_path: Path, root: Path, scanned_count: int, skipped_count: int) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Secrets"

    headers = [
        "Severity", "Confidence", "Rule", "File", "Line", "Column",
        "Secret Masked", "Secret SHA256 Prefix", "Entropy", "Likely Dummy/Test Value",
        "Context", "Matched Text", "Description", "Tags"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    severity_fills = {
        "Critical": PatternFill("solid", fgColor="F4CCCC"),
        "High": PatternFill("solid", fgColor="FCE5CD"),
        "Medium": PatternFill("solid", fgColor="FFF2CC"),
        "Low": PatternFill("solid", fgColor="D9EAD3"),
    }

    for f in findings:
        ws.append([
            f.severity, f.confidence, f.rule_name, f.file, f.line, f.column,
            f.secret_masked, f.secret_sha256_12, f.entropy, "Yes" if f.likely_dummy else "No",
            f.context, f.matched_text, f.description, f.tags,
        ])
        row = ws.max_row
        fill = severity_fills.get(f.severity)
        if fill:
            ws.cell(row=row, column=1).fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    widths = {
        1: 12, 2: 12, 3: 34, 4: 60, 5: 8, 6: 8, 7: 36, 8: 18,
        9: 10, 10: 20, 11: 90, 12: 80, 13: 55, 14: 30,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ["Scan root", str(root)],
        ["Generated UTC", datetime.now(timezone.utc).isoformat()],
        ["Files scanned", scanned_count],
        ["Files skipped", skipped_count],
        ["Total findings", len(findings)],
    ]
    summary.append(["Metric", "Value"])
    for r in summary_rows:
        summary.append(r)
    summary.append([])
    summary.append(["Severity", "Count"])
    for sev in ["Critical", "High", "Medium", "Low"]:
        summary.append([sev, sum(1 for f in findings if f.severity == sev)])
    summary.append([])
    summary.append(["Confidence", "Count"])
    for conf in ["High", "Medium", "Low"]:
        summary.append([conf, sum(1 for f in findings if f.confidence == conf)])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 90

    wb.save(out_path)


def write_csv(findings: list[Finding], out_path: Path) -> None:
    fieldnames = list(asdict(findings[0]).keys()) if findings else [
        "rule_name", "severity", "confidence", "file", "line", "column",
        "matched_text", "secret_masked", "secret_sha256_12", "entropy",
        "likely_dummy", "context", "description", "tags"
    ]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for finding in findings:
            writer.writerow(asdict(finding))


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Recursive hardcoded secret scanner for Java Spring/Struts source code.")
    parser.add_argument("root", help="Repository/source root to scan")
    parser.add_argument("-o", "--output", default="hardcoded_secrets.xlsx", help="Excel output path")
    parser.add_argument("--csv", dest="csv_output", help="Optional CSV output path")
    parser.add_argument("--include-tests", action="store_true", help="Include tests, mocks, samples, fixtures, and examples")
    parser.add_argument("--max-file-size", type=int, default=MAX_DEFAULT_FILE_SIZE_BYTES, help="Max file size in bytes to scan")
    parser.add_argument("--exclude-dir", action="append", default=[], help="Additional directory name to exclude; can be repeated")
    parser.add_argument("--include-ext", action="append", default=[], help="Additional extension to include, e.g. .txt; can be repeated")
    parser.add_argument("--fail-on-findings", action="store_true", help="Exit with code 2 if findings are detected")
    args = parser.parse_args(argv)

    root = Path(args.root).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        print(f"ERROR: root does not exist or is not a directory: {root}", file=sys.stderr)
        return 1

    extra_excluded_dirs = set(args.exclude_dir or [])
    extra_included_exts = {e if e.startswith(".") else f".{e}" for e in (args.include_ext or [])}

    findings: list[Finding] = []
    scanned_count = 0
    skipped_count = 0

    # Count skipped approximately while walking. The actual scanner only opens allowed text-like files.
    for current_root, dirs, files in os.walk(root):
        for filename in files:
            path = Path(current_root) / filename
            skip, _ = should_skip_file(path, root, args.include_tests, args.max_file_size, extra_excluded_dirs, extra_included_exts)
            if skip:
                skipped_count += 1

    for path in iter_files(root, args.include_tests, args.max_file_size, extra_excluded_dirs, extra_included_exts):
        text = read_text_safely(path)
        if text is None:
            skipped_count += 1
            continue
        scanned_count += 1
        findings.extend(scan_text(path, root, text))

    findings = sort_findings(findings)
    output = Path(args.output).expanduser().resolve()
    write_excel(findings, output, root, scanned_count, skipped_count)

    if args.csv_output:
        write_csv(findings, Path(args.csv_output).expanduser().resolve())

    print(f"Scan complete. Files scanned: {scanned_count}. Findings: {len(findings)}")
    print(f"Excel report: {output}")
    if args.csv_output:
        print(f"CSV report: {Path(args.csv_output).expanduser().resolve()}")

    if args.fail_on_findings and findings:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
